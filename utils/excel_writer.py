"""
Excel 写入工具模块
用于将数据导出到 Excel 文件
"""

import os
from typing import List, Dict
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import logging

logger = logging.getLogger(__name__)


class ExcelWriter:
    """Excel 文件写入器"""

    def __init__(self, output_dir: str = "data", sheet_name: str = "Datasets"):
        """
        初始化 Excel 写入器

        Args:
            output_dir: 输出目录
            sheet_name: 工作表名称
        """
        self.output_dir = output_dir
        self.sheet_name = sheet_name

        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)

        logger.info(f"初始化 Excel 写入器，输出目录: {output_dir}")

    def write_to_excel(self, data: List[Dict], filename: str,
                       columns: List[str] = None) -> str:
        """
        将数据写入 Excel 文件

        Args:
            data: 数据列表
            filename: 文件名
            columns: 列名列表（可选）

        Returns:
            输出文件的完整路径
        """
        if not data:
            logger.warning("没有数据需要写入")
            return None

        # 构建完整文件路径
        filepath = os.path.join(self.output_dir, filename)

        try:
            # 使用 pandas 创建 DataFrame
            df = pd.DataFrame(data)

            # 如果指定了列名，按指定顺序排列
            if columns:
                # 只保留存在的列
                existing_columns = [col for col in columns if col in df.columns]
                df = df[existing_columns]

            # 写入 Excel 文件
            df.to_excel(filepath, sheet_name=self.sheet_name, index=False, engine='openpyxl')

            logger.info(f"成功写入 {len(data)} 条数据到 {filepath}")

            # 美化 Excel 格式
            self._format_excel(filepath)

            return filepath

        except Exception as e:
            logger.error(f"写入 Excel 文件失败: {e}")
            raise

    def _format_excel(self, filepath: str):
        """
        美化 Excel 文件格式

        Args:
            filepath: Excel 文件路径
        """
        try:
            from openpyxl import load_workbook

            # 加载工作簿
            wb = load_workbook(filepath)
            ws = wb.active

            # 设置标题行样式
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            # 应用标题行样式
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 自动调整列宽
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter

                for cell in column:
                    try:
                        if cell.value:
                            # 计算字符串长度（中文算2个字符）
                            length = self._get_display_width(str(cell.value))
                            if length > max_length:
                                max_length = length
                    except:
                        pass

                # 设置列宽（最小10，最大50）
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # 设置对齐方式
            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # 保存修改
            wb.save(filepath)
            logger.info(f"Excel 格式美化完成")

        except Exception as e:
            logger.warning(f"美化 Excel 格式失败: {e}")

    @staticmethod
    def _get_display_width(text: str) -> int:
        """
        计算文本显示宽度（中文字符算2个宽度）

        Args:
            text: 文本

        Returns:
            显示宽度
        """
        width = 0
        for char in text:
            if ord(char) > 127:  # 非ASCII字符（如中文）
                width += 2
            else:
                width += 1
        return width

    def append_to_excel(self, data: List[Dict], filename: str,
                       columns: List[str] = None) -> str:
        """
        追加数据到现有 Excel 文件

        Args:
            data: 数据列表
            filename: 文件名
            columns: 列名列表

        Returns:
            输出文件的完整路径
        """
        filepath = os.path.join(self.output_dir, filename)

        try:
            # 检查文件是否存在
            if os.path.exists(filepath):
                # 读取现有数据
                existing_df = pd.read_excel(filepath, engine='openpyxl')
                new_df = pd.DataFrame(data)

                # 合并数据
                combined_df = pd.concat([existing_df, new_df], ignore_index=True)

                # 写回文件
                combined_df.to_excel(filepath, sheet_name=self.sheet_name,
                                    index=False, engine='openpyxl')

                logger.info(f"成功追加 {len(data)} 条数据到 {filepath}")
            else:
                # 文件不存在，创建新文件
                return self.write_to_excel(data, filename, columns)

            return filepath

        except Exception as e:
            logger.error(f"追加数据到 Excel 失败: {e}")
            raise

    def write_multiple_sheets(self, data_dict: Dict[str, List[Dict]], filename: str):
        """
        写入多个工作表

        Args:
            data_dict: 字典，键为工作表名，值为数据列表
            filename: 文件名

        Returns:
            输出文件的完整路径
        """
        filepath = os.path.join(self.output_dir, filename)

        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                for sheet_name, data in data_dict.items():
                    if data:
                        df = pd.DataFrame(data)
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

            logger.info(f"成功写入多个工作表到 {filepath}")
            return filepath

        except Exception as e:
            logger.error(f"写入多个工作表失败: {e}")
            raise
